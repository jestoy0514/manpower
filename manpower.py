#!/usr/bin/python3

import tkinter as tk
from tkinter import ttk
from tkinter import messagebox as mb
from tkinter import filedialog as fd
from tkinter.scrolledtext import ScrolledText
import sys
import os
from fpdf import FPDF
from PIL import Image, ImageTk
import csv
import glob
import subprocess
import calendar as cl
from openpyxl import Workbook
import webbrowser
from datetime import datetime
import time

from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker, join
from dtbase.dtbase import *


__version__ = '1.0.0'
__author__ = 'Jesus Vedasto Olazo'
__email__ = 'jestoy.olazo@gmail.com'
__license__ = 'MIT'

HOME_DIR = os.path.join(os.path.expanduser('~'), '.manpower_mngt')
DB_FILE = os.path.join(HOME_DIR, 'manpower_db.db')
ENGINE = create_engine(f'sqlite:///{DB_FILE}')

if not os.path.isdir(HOME_DIR):
    os.mkdir(HOME_DIR)

if not os.path.isfile(DB_FILE):
    Base.metadata.create_all(ENGINE)

Base.metadata.bind = ENGINE
DBSession = sessionmaker(bind=ENGINE)

def image_list(size=(16, 16)):
    """This creates a dictionary of image file path for software icons."""
    img_lst = {}
    for img in glob.glob('images/*.png'):
        path_to_file = img
        img = os.path.basename(img)
        file_name = img.split('.')[0]
        img_lst[file_name] = ImageTk.PhotoImage(Image.open(path_to_file).resize(size))
    return img_lst

# Start of MainWindow class
class MainWindow(ttk.Frame):
    def __init__(self, master=None, *args, **kwargs):
        super(MainWindow, self).__init__(master, *args, **kwargs)
        self.master.protocol('WM_DELETE_WINDOW', self.close_app)
        #self.master.geometry(f'{950}x{600}+{20}+{10}')
        self.master.title(f'Manpower Management {__version__}')
        self.img_list = image_list()
        self.img_list36 = image_list(size=(36, 36))
        if os.name == 'nt':
            # Set window icon if using Windows operating system.
            self.master.state("zoomed")
            self.master.iconbitmap("manpower.ico")
        elif os.name == 'posix':
            self.master.attributes("-zoomed", True)
            # Set window icon if using Linux operating system.
            self.master.tk.call('wm', 'iconphoto', self.master._w, self.img_list36['manpower'])
        self.setup_ui()

    def setup_ui(self):
        """ This module initialize all the remaining widgets. """
        # This is the menubar of the application.
        menubar = tk.Menu(self)
        self.master.config(menu=menubar)

        filemenu = tk.Menu(menubar, tearoff=0)
        helpmenu = tk.Menu(menubar, tearoff=0)

        menubar.add_cascade(label='File', menu=filemenu)
        menubar.add_cascade(label='Help', menu=helpmenu)

        filemenu.add_command(label='Project', image=self.img_list['project1'], compound=tk.LEFT, command=self.project_window)
        filemenu.add_command(label='Designation', image=self.img_list['contractor'], compound=tk.LEFT, command=self.designation_window)
        filemenu.add_command(label='Transaction', image=self.img_list['labors'], compound=tk.LEFT, command=self.designation_window)
        filemenu.add_separator()
        filemenu.add_command(label='Quit', image=self.img_list['quit'], compound=tk.LEFT, command=self.close_app)

        helpmenu.add_command(label='Help', image=self.img_list['help'], compound=tk.LEFT, command=self.show_help)
        helpmenu.add_separator()
        helpmenu.add_command(label='About', image=self.img_list['about'], compound=tk.LEFT, command=self.show_about)

        # All this frame are containers of other widgets.
        title_frame = ttk.Frame(self)
        title_frame.pack(fill=tk.X, padx=5, pady=5)
        search_frame = ttk.Frame(self)
        search_frame.pack(fill=tk.X, padx=5, pady=5)
        view_frame = ttk.Frame(self)
        view_frame.pack(expand=True, fill=tk.BOTH, padx=5, pady=5)
        btn_frame = ttk.Frame(self)
        btn_frame.pack(fill=tk.X, padx=5, pady=5)

        # All widgets for title_frame
        proj_btn = ttk.Button(title_frame, image=self.img_list36['project1'], command=self.project_window)
        proj_btn.pack(side=tk.LEFT)
        desig_btn = ttk.Button(title_frame, image=self.img_list36['contractor'], command=self.designation_window)
        desig_btn.pack(side=tk.LEFT)
        tr_btn = ttk.Button(title_frame, image=self.img_list36['labors'], command=self.transaction_window)
        tr_btn.pack(side=tk.LEFT)
        app_lbl = tk.Label(title_frame, text=f'Manpower\nManagement {__version__}', font='Times 19 bold italic', fg='#FF2100')
        app_lbl.pack(side=tk.RIGHT)

        # All widgets for search_frame
        date_lbl = ttk.Label(search_frame, text='Date:')
        date_lbl.pack(side=tk.LEFT)
        self.search_entry = ttk.Entry(search_frame)
        self.search_entry.pack(side=tk.LEFT)
        self.search_entry.insert(tk.END, datetime.strftime(datetime.now(), '%d/%m/%Y'))
        cal_btn = ttk.Button(search_frame, image=self.img_list['calendar'], command=self.change_date)
        cal_btn.pack(side=tk.LEFT)

        # All widgets for view_frame
        cols = ('project', 'present', 'absent', 'vacation', 'total', 'remarks')
        self.manp_view = ttk.Treeview(view_frame, columns=cols)
        self.manp_view.pack(expand=True, fill=tk.BOTH, side=tk.LEFT)
        self.manp_view.heading('#0', text='S. No.')
        self.manp_view.column('#0', width=60, stretch=False, anchor=tk.CENTER)
        self.manp_view.bind('<<TreeviewSelect>>', self.load_details)

        for col in cols:
            self.manp_view.heading(col, text=col.upper())
            if col in cols[1:5]:
                self.manp_view.column(col, width=100, stretch=False, anchor=tk.CENTER)

        vbar = ttk.Scrollbar(view_frame, orient=tk.VERTICAL)
        vbar.pack(side=tk.LEFT, fill=tk.Y)

        vbar.config(command=self.manp_view.yview)
        self.manp_view['yscrollcommand'] = vbar.set

        self.manp_view.tag_configure("odd", background="#d5f4e6", font=('Times', 11, ''))
        self.manp_view.tag_configure("even", background="#80ced6", font=('Times', 11, ''))
        self.manp_view.tag_configure("total_color", background="#1C1C7B", foreground='white', font=('Times', 11, 'bold italic'))

        # Widgets for details_frame
        details_frame = ttk.Labelframe(view_frame, text='Details')
        details_frame.pack(side=tk.RIGHT, fill=tk.Y)
        self.dets_view = ttk.Treeview(details_frame)
        self.dets_view.pack(expand=True, fill=tk.BOTH, padx=5, pady=5)

        self.dets_view.tag_configure("odd", background="#d5f4e6", font=('Times', 11, ''))
        self.dets_view.tag_configure("even", background="#80ced6", font=('Times', 11, ''))
        self.dets_view.tag_configure("total_color", background="#1C1C7B", foreground='white', font=('Times', 11, 'bold italic'))

        dets_cols = ('designation', 'pre', 'abs', 'vac', 'total')
        self.dets_view['columns'] = dets_cols

        self.dets_view.column('#0', width=40, stretch=False, anchor=tk.CENTER)
        self.dets_view.heading('#0', text='ID')

        for dets_col in dets_cols:
            if dets_col != 'designation':
                self.dets_view.column(dets_col, width=50, stretch=False, anchor=tk.CENTER)
            self.dets_view.heading(dets_col, text=dets_col.upper())


        # All widgets for btn_frame
        export_btn = ttk.Button(btn_frame, text='Export', image=self.img_list['export'], compound=tk.LEFT, command=self.export_records)
        export_btn.pack(side=tk.LEFT)
        print_btn = ttk.Button(btn_frame, text='Print', image=self.img_list['printer'], compound=tk.LEFT, command=self.print_record)
        print_btn.pack(side=tk.LEFT)
        close_btn = ttk.Button(btn_frame, text='Close', image=self.img_list['cancel'], compound=tk.LEFT, command=self.close_app)
        close_btn.pack(side=tk.RIGHT)

        self.update_view()
        self.search_entry.focus_set()

    def load_details(self, event):
        children = self.dets_view.get_children()
        self.dets_view.delete(*children)
        tr_date = datetime.strptime(self.search_entry.get(), '%d/%m/%Y')
        item = self.manp_view.focus()
        session = DBSession()
        if item != '':
            proj = self.manp_view.item(item)['values'][0]
            proj_rec = session.query(Project).filter(Project.name == proj).first()
            trans_recs = session.query(TransactionDetails).join(Transaction).join(Designation).filter(Transaction.project == proj_rec).filter(Transaction.tr_date == tr_date).all()
            sum_pre = 0
            sum_abs = 0
            sum_vac = 0
            counter = 1
            for trans_rec in trans_recs:
                values = (trans_rec.designation.name, f'{trans_rec.present}', f'{trans_rec.absent}', f'{trans_rec.vacation}', f'{trans_rec.present+trans_rec.absent+trans_rec.vacation}')
                if counter % 2 == 0:
                    self.dets_view.insert('', tk.END, f'{counter}', text=f'{counter}', values=values, tags='even')
                else:
                    self.dets_view.insert('', tk.END, f'{counter}', text=f'{counter}', values=values, tags='odd')
                sum_pre+=trans_rec.present
                sum_abs+=trans_rec.absent
                sum_vac+=trans_rec.vacation
                counter+=1
            values = ('Total', f'{sum_pre}', f'{sum_abs}', f'{sum_vac}', f'{sum_pre+sum_abs+sum_vac}')
            self.dets_view.insert('', tk.END, 'total', text='', values=values, tags='total_color')
        session.close()

    def update_view(self):
        children = self.manp_view.get_children()
        self.manp_view.delete(*children)
        session = DBSession()
        test_dict = {}
        date_today = datetime.strptime(self.search_entry.get(), '%d/%m/%Y')

        try:
            proj_recs = session.query(Project).all()
            for proj_rec in proj_recs:
                trans_recs = session.query(Transaction).filter(Transaction.project == proj_rec).filter(Transaction.tr_date == date_today).all()
                sum_present = 0
                sum_absent = 0
                sum_vacation = 0
                remarks = ''
                for trans_rec in trans_recs:
                    details_recs = session.query(TransactionDetails).filter(TransactionDetails.transaction == trans_rec).all()
                    remarks = trans_rec.remarks
                    for details_rec in details_recs:
                        #print(details_rec.present, details_rec.absent, details_rec.vacation)
                        sum_present+=details_rec.present
                        sum_absent+=details_rec.absent
                        sum_vacation+=details_rec.vacation
                if (sum_present+sum_absent+sum_vacation) != 0:
                    test_dict[proj_rec.name] = (sum_present, sum_absent, sum_vacation, remarks)
                else:
                    if remarks != '':
                        test_dict[proj_rec.name] = (sum_present, sum_absent, sum_vacation, remarks)
            #print(test_dict)
            counter = 1
            pre_tl = 0
            abs_tl = 0
            vac_tl = 0
            for k, v in test_dict.items():
                values = (k, f'{v[0]}', f'{v[1]}', f'{v[2]}', f'{v[0]+v[1]+v[2]}', v[3] )
                if counter % 2 == 0:
                    self.manp_view.insert('', tk.END, f'{counter}', text=f'{counter}', tags='even', values=values)
                else:
                    self.manp_view.insert('', tk.END, f'{counter}', text=f'{counter}', tags='odd', values=values)
                counter+=1
                pre_tl+=v[0]
                abs_tl+=v[1]
                vac_tl+=v[2]
            if len(test_dict) != 0:
                values = ('Total', f'{pre_tl}', f'{abs_tl}', f'{vac_tl}', f'{pre_tl+abs_tl+vac_tl}', '' )
                self.manp_view.insert('', tk.END, 'total', text='', tags='total_color', values=values)
        except:
            print(sys.exc_info())
        finally:
            session.close()

    def project_window(self):
        tp = tk.Toplevel(self)
        win = ProjectWindow(tp)
        win.pack(expand=True, fill=tk.BOTH, padx=5, pady=5)

    def designation_window(self):
        tp = tk.Toplevel(self)
        win = DesignationWindow(tp)
        win.pack(expand=True, fill=tk.BOTH, padx=5, pady=5)

    def transaction_window(self):
        tp = tk.Toplevel(self)
        win = TransactionWindow(tp)
        win.pack(expand=True, fill=tk.BOTH, padx=5, pady=5)
        self.wait_window(tp)
        self.update_view()

    def show_help(self):
        webbrowser.open_new_tab('index.html')

    def show_about(self):
        tp = tk.Toplevel(self)
        win = AboutWindow(tp)
        win.pack(expand=True, fill=tk.BOTH, padx=2, pady=2)

    def change_date(self):
        date_now = datetime.today()
        tp = tk.Toplevel(self)
        win = CalendarWidget(date_now.year, date_now.month, tp)
        win.pack(expand=True, fill=tk.BOTH, padx=10, pady=10)
        self.wait_window(tp)
        if win.date is not None:
            self.search_entry.delete('0', tk.END)
            self.search_entry.insert(tk.END, win.date)
            self.update_view()

    def export_records(self):
        session = DBSession()
        test_dict = {}
        date_today = datetime.strptime(self.search_entry.get(), '%d/%m/%Y')

        proj_recs = session.query(Project).all()
        for proj_rec in proj_recs:
            trans_recs = session.query(Transaction).filter(Transaction.project == proj_rec).filter(Transaction.tr_date == date_today).all()
            sum_present = 0
            sum_absent = 0
            sum_vacation = 0
            remarks = ''
            for trans_rec in trans_recs:
                details_recs = session.query(TransactionDetails).filter(TransactionDetails.transaction == trans_rec).all()
                remarks = trans_rec.remarks
                for details_rec in details_recs:
                    #print(details_rec.present, details_rec.absent, details_rec.vacation)
                    sum_present+=details_rec.present
                    sum_absent+=details_rec.absent
                    sum_vacation+=details_rec.vacation
            if (sum_present+sum_absent+sum_vacation) != 0:
                test_dict[proj_rec.name] = (sum_present, sum_absent, sum_vacation, remarks)
            else:
                if remarks != '':
                    test_dict[proj_rec.name] = (sum_present, sum_absent, sum_vacation, remarks)

        filename = fd.asksaveasfilename(parent=self, defaultextension='.xlsx', initialfile=f'Manpower-{self.search_entry.get().replace("/", "-")}.xlsx')

        if filename == '':
            session.close()
            return

        wb = Workbook()
        ws = wb.active
        heading = ('S. No.', 'Project', 'Present', 'Absent', 'Vacation', 'Total', 'Remarks')
        letter = ('A', 'B', 'C', 'D', 'E', 'F', 'G')
        for idx, value in enumerate(heading):
            ws[f'{letter[idx]}1'] = value

        counter = 2
        for k, v in test_dict.items():
            ws[f'A{counter}'] = counter - 1
            ws[f'B{counter}'] = k
            ws[f'C{counter}'] = v[0]
            ws[f'D{counter}'] = v[1]
            ws[f'E{counter}'] = v[2]
            total = v[0] + v[1] + v[2]
            ws[f'F{counter}'] = total
            ws[f'G{counter}'] = v[3]
            counter+=1
        session.close()
        wb.save(filename)
        time.sleep(1)
        if os.name == 'nt':
            CREATE_NO_WINDOW = 0x08000000
            subprocess.call(["start", filename], creationflags=CREATE_NO_WINDOW, shell=True)
        else:
            os.system("xdg-open %s"% (filename,))

    def print_record(self):
        session = DBSession()
        test_dict = {}
        date_today = datetime.strptime(self.search_entry.get(), '%d/%m/%Y')

        try:
            proj_recs = session.query(Project).all()
            for proj_rec in proj_recs:
                trans_recs = session.query(Transaction).filter(Transaction.project == proj_rec).filter(Transaction.tr_date == date_today).all()
                sum_present = 0
                sum_absent = 0
                sum_vacation = 0
                remarks = ''
                for trans_rec in trans_recs:
                    details_recs = session.query(TransactionDetails).filter(TransactionDetails.transaction == trans_rec).all()
                    remarks = trans_rec.remarks
                    for details_rec in details_recs:
                        #print(details_rec.present, details_rec.absent, details_rec.vacation)
                        sum_present+=details_rec.present
                        sum_absent+=details_rec.absent
                        sum_vacation+=details_rec.vacation
                if (sum_present+sum_absent+sum_vacation) != 0:
                    test_dict[proj_rec.name] = (sum_present, sum_absent, sum_vacation, remarks)
                else:
                    if remarks != '':
                        test_dict[proj_rec.name] = (sum_present, sum_absent, sum_vacation, remarks)

            if len(test_dict) != 0:
                filename = fd.asksaveasfilename(parent=self, defaultextension='.pdf',
                                        initialfile=f'ManpowerSummary-{self.search_entry.get().replace("/", "")}')
                if filename != '':
                    pdf = PDF('L')
                    pdf.alias_nb_pages()
                    pdf.add_page()
                    pdf.set_font('Times', '', 12)
                    pdf.cell(0, 10, 'Date: '+self.search_entry.get(), 0, 1)
                    pdf.set_font('Times', 'BI', 12)
                    pdf.set_fill_color(86, 159, 169)
                    pdf.set_text_color(255, 255, 255)
                    pdf.cell(20, 10, 'Sl. No.', 1, 0, 'C', True)
                    pdf.cell(70, 10, 'Project', 1, 0, 'C', True)
                    pdf.cell(28, 10, 'Present', 1, 0, 'C', True)
                    pdf.cell(28, 10, 'Absent', 1, 0, 'C', True)
                    pdf.cell(28, 10, 'Vacation', 1, 0, 'C', True)
                    pdf.cell(28, 10, 'Total', 1, 0, 'C', True)
                    pdf.cell(0, 10, 'Remarks', 1, 1, 'C', True)
                    pdf.set_font('Times', '', 12)
                    pdf.set_text_color(0, 0, 0)
                    counter = 1
                    sum_pre = 0
                    sum_abs = 0
                    sum_vac = 0
                    for k,v in test_dict.items():
                        if counter % 2 == 0:
                            pdf.set_fill_color(152, 211, 219)
                        else:
                            pdf.set_fill_color(195, 223, 227)
                        pdf.cell(20, 8, f'{counter}', 1, 0, 'C', True)
                        pdf.cell(70, 8, k, 1, 0, '', True)
                        pdf.cell(28, 8, f'{v[0]}', 1, 0, 'C', True)
                        pdf.cell(28, 8, f'{v[1]}', 1, 0, 'C', True)
                        pdf.cell(28, 8, f'{v[2]}', 1, 0, 'C', True)
                        total = v[0] + v[1] + v[2]
                        pdf.cell(28, 8, f'{total}', 1, 0, 'C', True)
                        pdf.cell(0, 8, v[3], 1, 1, 'C', True)
                        sum_pre+=v[0]
                        sum_abs+=v[1]
                        sum_vac+=v[2]
                        counter+=1
                    pdf.set_font('Times', 'BI', 12)
                    if counter % 2 == 0:
                        pdf.set_fill_color(152, 211, 219)
                    else:
                        pdf.set_fill_color(195, 223, 227)
                    pdf.cell(20, 10, '', 1, 0, 'C', True)
                    pdf.cell(70, 10, 'Total', 1, 0, 'C', True)
                    pdf.cell(28, 10, f'{sum_pre}', 1, 0, 'C', True)
                    pdf.cell(28, 10, f'{sum_abs}', 1, 0, 'C', True)
                    pdf.cell(28, 10, f'{sum_vac}', 1, 0, 'C', True)
                    pdf.cell(28, 10, f'{sum_pre+sum_abs+sum_vac}', 1, 0, 'C', True)
                    pdf.cell(0, 10, '', 1, 1, 'C', True)
                    pdf.output(filename, 'F')
                    if os.name == 'nt':
                        CREATE_NO_WINDOW = 0x08000000
                        subprocess.call(["start", filename], creationflags=CREATE_NO_WINDOW, shell=True)
                    else:
                        os.system(f"xdg-open {filename}")
        except:
            print(sys.exc_info())
        finally:
            session.close()

    def close_app(self):
        self.master.destroy()
# End of MainWindow class

# Start of PDF class
class PDF(FPDF):
    def header(self):
        self.set_font('Times', 'B', 20)
        self.cell(0, 10, 'Al Hamra Construction Co. LLC', 0, 1, 'C')
        self.set_font('Times', 'I', 17)
        self.cell(0, 20, 'Daily Manpower Report', 0, 1, 'C')

    def footer(self):
        self.set_y(-15)
        self.set_font('Arial', 'I', 8)
        self.cell(0, 10, 'Page ' + str(self.page_no()) + ' of {nb}', 0, 0, 'C')
# End of PDF class

# Start of ProjectWindow class
class ProjectWindow(tk.Frame):
    def __init__(self, master=None, *args, **kwargs):
        tk.Frame.__init__(self, master, *args, **kwargs)
        self.master.protocol('WM_DELETE_WINDOW', self.close_app)
        self.master.title('Projects')
        # self.master.geometry('600x400+20+20')
        self.img_list = image_list()
        self.img_list36 = image_list(size=(36, 36))
        if os.name == "nt":
            self.master.iconbitmap('manpower.ico')
        elif os.name == 'posix':
            self.master.tk.call('wm', 'iconphoto', self.master._w, self.img_list36['manpower'])
        self.record_id = None
        self.style = ttk.Style()
        self.setup_ui()

    def setup_ui(self):
        top_frame = ttk.Frame(self)
        top_frame.pack(fill=tk.X, padx=5, pady=5)
        mid_frame = ttk.Frame(self)
        mid_frame.pack(expand=True, fill=tk.BOTH, padx=5, pady=5)
        bot_frame = ttk.Frame(self)
        bot_frame.pack(fill=tk.X, padx=5, pady=5)

        proj_name = ttk.Label(top_frame, text='Project Name:')
        proj_name.pack(side=tk.LEFT, padx=5, pady=5)
        self.proj_entry = ttk.Entry(top_frame)
        self.proj_entry.pack(side=tk.LEFT, padx=5, pady=5)
        self.proj_entry.config(font=('Cambria', 10, 'bold'))
        self.save_btn = ttk.Button(top_frame, text='Save', image=self.img_list['save'], compound=tk.LEFT)
        self.save_btn.pack(side=tk.LEFT, padx=5, pady=5)
        self.save_btn.config(command=self.save_record)
        self.search_btn = ttk.Button(top_frame, text='Search', image=self.img_list['search'], compound=tk.LEFT)
        self.search_btn.pack(side=tk.LEFT, padx=5, pady=5)
        self.search_btn.config(command=self.search_record)

        self.proj_view = ttk.Treeview(mid_frame)
        self.proj_view.pack(expand=True, fill=tk.BOTH, side=tk.LEFT)
        self.proj_view['columns'] = ('project',)
        self.proj_view.heading('#0', text='ID')
        self.proj_view.column('#0', width=50, stretch=False)
        self.proj_view.heading('project', text='PROJECT')

        self.proj_view.tag_configure("odd", background="#d5f4e6", font=('Times', 11, ''))
        self.proj_view.tag_configure("even", background="#80ced6", font=('Times', 11, ''))

        self.vbar = ttk.Scrollbar(mid_frame, orient=tk.VERTICAL)
        self.vbar.pack(side=tk.RIGHT, fill=tk.Y)

        self.vbar.config(command=self.proj_view.yview)
        self.proj_view['yscrollcommand'] = self.vbar.set

        self.update_view()

        self.edit_btn = ttk.Button(bot_frame, text='Edit', image=self.img_list['edit'], compound=tk.LEFT)
        self.edit_btn.pack(side=tk.LEFT, padx=5, pady=5)
        self.edit_btn.config(command=self.edit_record)
        self.delete_btn = ttk.Button(bot_frame, text='Delete', image=self.img_list['minus'], compound=tk.LEFT)
        self.delete_btn.pack(side=tk.LEFT, padx=5, pady=5)
        self.delete_btn.config(command=self.delete_record)
        self.export_btn = ttk.Button(bot_frame, text='Export', image=self.img_list['export'], compound=tk.LEFT)
        self.export_btn.pack(side=tk.LEFT, padx=5, pady=5)
        self.export_btn.config(command=self.export_record)
        self.close_btn = ttk.Button(bot_frame, text='Close', image=self.img_list['cancel'], compound=tk.LEFT)
        self.close_btn.pack(side=tk.RIGHT, padx=5, pady=5)
        self.close_btn.config(command=self.close_app)

        self.proj_entry.focus_set()

    def update_view(self):
        session = DBSession()
        records = session.query(Project).all()
        children = self.proj_view.get_children()
        if len(children) != 0:
            for child in children:
                self.proj_view.delete(child)
        if not len(records) == 0:
            counter = 1
            for record in records:
                if counter % 2 == 0:
                    self.proj_view.insert('', tk.END, str(record.id), text=str(record.id), tags='even')
                else:
                    self.proj_view.insert('', tk.END, str(record.id), text=str(record.id), tags='odd')
                self.proj_view.set(str(record.id), 'project', str(record.name))
                counter+=1
        session.close()

    def save_record(self):
        project_name = self.proj_entry.get()
        if self.record_id != None:
            session = DBSession()
            record = session.query(Project).filter(Project.id == self.record_id).first()
            record.name = project_name
            session.commit()
            session.close()
            self.record_id = None
        else:
            if project_name != '':
                session = DBSession()
                new_record = Project(name=project_name)
                session.add(new_record)
                session.commit()
                session.close()
        self.update_view()
        self.proj_entry.delete(0, tk.END)
        self.proj_entry.focus_set()

    def search_record(self):
        keyword = self.proj_entry.get()
        if keyword == 'all':
            self.update_view()
            return
        if keyword != '':
            children = self.proj_view.get_children()
            if len(children) != 0:
                for child in children:
                    self.proj_view.delete(child)
            session = DBSession()
            records = session.query(Project).filter(Project.name.like(f'%{keyword}%')).all()
            if len(records) != 0:
                counter = 1
                for record in records:
                    if counter % 2 == 0:
                        self.proj_view.insert('', tk.END, str(record.id), text=str(record.id), tags='even')
                    else:
                        self.proj_view.insert('', tk.END, str(record.id), text=str(record.id), tags='odd')
                    self.proj_view.set(str(record.id), 'project', str(record.name))
                    counter+=1
            else:
                self.update_view()
                mb.showwarning('No Records Found', 'Sorry no record(s) has been found', parent=self)

    def edit_record(self):
        record = self.proj_view.focus()
        if record != '':
            self.record_id = int(record)
            self.proj_entry.delete('0', tk.END)

            item = self.proj_view.item(record, 'values')
            self.proj_entry.insert(tk.END, item[0])
        else:
            mb.showwarning('No Record', 'Please select a record and try again.', parent=self)

    def delete_record(self):
        '''This method delete the highlighted record from the database.'''
        record = self.proj_view.focus()
        if record != '':
            # Checks if a record has been selected
            record_id = int(record)
            session = DBSession()
            # Query the record from database.
            record = session.query(Project).filter(Project.id == record_id).first()
            # This line delete the record.
            session.delete(record)
            # Applied all the pending transaction to the database.
            session.commit()
            # Explicitly close database connection.
            session.close()
            # Reload the projects view and display it.
            self.update_view()
        else:
            # Trow a messagebox if no record is selected.
            mb.showwarning('No Record', 'Please select a record and try again.', parent=self)

    def export_record(self):
        '''This method is use to export project table to excel file.'''
        # Ask the location to save the file.
        filename = fd.asksaveasfilename(parent=self, defaultextension='.xls', initialfile='project_details.xls')
        # Checks if a location has been selected, if not will skip the code below.
        if filename != '':
            # This create a temporary workbook and worksheet
            wb = Workbook()
            ws = wb.active

            session = DBSession()
            # Gets the details of project table
            records = session.query(Project).all()
            # If table is not empty begin writing it into excel worksheet.
            if len(records) != 0:
                ws['A1'] = 'ID'
                ws['B1'] = 'Project'
                counter = 2
                for record in records:
                    ws[f'A{counter}'] = str(record.id)
                    ws[f'B{counter}'] = str(record.name)
                    counter+=1
            # This will save the workbook from memory into a file.
            wb.save(filename)
            # Then finally closes the connection to the database.
            session.close()

    def close_app(self):
        self.master.destroy()
# End of ProjectWindow class

# Start of DesignationWindow class
class DesignationWindow(tk.Frame):
    def __init__(self, master=None, *args, **kwargs):
        super(DesignationWindow, self).__init__(master, *args, **kwargs)
        self.master.protocol('WM_DELETE_WINDOW', self.close_app)
        self.master.title('Designation')
        # self.master.geometry('600x400+20+20')
        self.img_list = image_list()
        self.img_list36 = image_list(size=(36, 36))
        if os.name == "nt":
            self.master.iconbitmap('manpower.ico')
        elif os.name == 'posix':
            self.master.tk.call('wm', 'iconphoto', self.master._w, self.img_list36['manpower'])
        self.record_id = None
        self.style = ttk.Style()
        self.setup_ui()

    def setup_ui(self):
        top_frame = ttk.Frame(self)
        top_frame.pack(fill=tk.X, padx=5, pady=5)
        mid_frame = ttk.Frame(self)
        mid_frame.pack(expand=True, fill=tk.BOTH, padx=5, pady=5)
        bot_frame = ttk.Frame(self)
        bot_frame.pack(fill=tk.X, padx=5, pady=5)

        proj_name = ttk.Label(top_frame, text='Designation:')
        proj_name.pack(side=tk.LEFT, padx=5, pady=5)
        self.proj_entry = ttk.Entry(top_frame)
        self.proj_entry.pack(side=tk.LEFT, padx=5, pady=5)
        self.proj_entry.config(font=('Cambria', 10, 'bold'))
        self.save_btn = ttk.Button(top_frame, text='Save', image=self.img_list['save'], compound=tk.LEFT)
        self.save_btn.pack(side=tk.LEFT, padx=5, pady=5)
        self.save_btn.config(command=self.save_record)
        self.search_btn = ttk.Button(top_frame, text='Search', image=self.img_list['search'], compound=tk.LEFT)
        self.search_btn.pack(side=tk.LEFT, padx=5, pady=5)
        self.search_btn.config(command=self.search_record)

        self.proj_view = ttk.Treeview(mid_frame)
        self.proj_view.pack(expand=True, fill=tk.BOTH, side=tk.LEFT)
        self.proj_view['columns'] = ('designation',)
        self.proj_view.heading('#0', text='ID')
        self.proj_view.column('#0', width=50, stretch=False)
        self.proj_view.heading('designation', text='DESIGNATION')

        self.proj_view.tag_configure("odd", background="#d5f4e6", font=('Times', 11, ''))
        self.proj_view.tag_configure("even", background="#80ced6", font=('Times', 11, ''))

        self.vbar = ttk.Scrollbar(mid_frame, orient=tk.VERTICAL)
        self.vbar.pack(side=tk.RIGHT, fill=tk.Y)

        self.vbar.config(command=self.proj_view.yview)
        self.proj_view['yscrollcommand'] = self.vbar.set

        self.update_view()

        self.edit_btn = ttk.Button(bot_frame, text='Edit', image=self.img_list['edit'], compound=tk.LEFT)
        self.edit_btn.pack(side=tk.LEFT, padx=5, pady=5)
        self.edit_btn.config(command=self.edit_record)
        self.delete_btn = ttk.Button(bot_frame, text='Delete', image=self.img_list['minus'], compound=tk.LEFT)
        self.delete_btn.pack(side=tk.LEFT, padx=5, pady=5)
        self.delete_btn.config(command=self.delete_record)
        self.export_btn = ttk.Button(bot_frame, text='Export', image=self.img_list['export'], compound=tk.LEFT)
        self.export_btn.pack(side=tk.LEFT, padx=5, pady=5)
        self.export_btn.config(command=self.export_record)
        self.close_btn = ttk.Button(bot_frame, text='Close', image=self.img_list['cancel'], compound=tk.LEFT)
        self.close_btn.pack(side=tk.RIGHT, padx=5, pady=5)
        self.close_btn.config(command=self.close_app)

        self.proj_entry.focus_set()

    def update_view(self):
        session = DBSession()
        records = session.query(Designation).all()
        children = self.proj_view.get_children()
        if len(children) != 0:
            for child in children:
                self.proj_view.delete(child)
        if not len(records) == 0:
            counter = 1
            for record in records:
                if counter % 2 == 0:
                    self.proj_view.insert('', tk.END, str(record.id), text=str(record.id), tags='even')
                else:
                    self.proj_view.insert('', tk.END, str(record.id), text=str(record.id), tags='odd')
                self.proj_view.set(str(record.id), 'designation', str(record.name))
                counter+=1
        session.close()

    def save_record(self):
        project_name = self.proj_entry.get()
        if self.record_id != None:
            session = DBSession()
            record = session.query(Designation).filter(Designation.id == self.record_id).first()
            record.name = project_name
            session.commit()
            session.close()
            self.record_id = None
        else:
            if project_name != '':
                session = DBSession()
                new_record = Designation(name=project_name)
                session.add(new_record)
                session.commit()
                session.close()
        self.update_view()
        self.proj_entry.delete(0, tk.END)
        self.proj_entry.focus_set()

    def search_record(self):
        keyword = self.proj_entry.get()
        if keyword == 'all':
            self.update_view()
            return
        if keyword != '':
            children = self.proj_view.get_children()
            if len(children) != 0:
                for child in children:
                    self.proj_view.delete(child)
            session = DBSession()
            records = session.query(Designation).filter(Designation.name.like(f'%{keyword}%')).all()
            if len(records) != 0:
                counter = 1
                for record in records:
                    if counter % 2 == 0:
                        self.proj_view.insert('', tk.END, str(record.id), text=str(record.id), tags='even')
                    else:
                        self.proj_view.insert('', tk.END, str(record.id), text=str(record.id), tags='odd')
                    self.proj_view.set(str(record.id), 'project', str(record.name))
                    counter+=1
            else:
                self.update_view()
                mb.showwarning('No Records Found', 'Sorry no record(s) has been found', parent=self)

    def edit_record(self):
        record = self.proj_view.focus()
        if record != '':
            self.record_id = int(record)
            self.proj_entry.delete('0', tk.END)

            item = self.proj_view.item(record, 'values')
            self.proj_entry.insert(tk.END, item[0])
        else:
            mb.showwarning('No Record', 'Please select a record and try again.', parent=self)

    def delete_record(self):
        '''This method delete the highlighted record from the database.'''
        record = self.proj_view.focus()
        if record != '':
            # Checks if a record has been selected
            record_id = int(record)
            session = DBSession()
            # Query the record from database.
            record = session.query(Designation).filter(Designation.id == record_id).first()
            # This line delete the record.
            session.delete(record)
            # Applied all the pending transaction to the database.
            session.commit()
            # Explicitly close database connection.
            session.close()
            # Reload the projects view and display it.
            self.update_view()
        else:
            # Trow a messagebox if no record is selected.
            mb.showwarning('No Record', 'Please select a record and try again.', parent=self)

    def export_record(self):
        '''This method is use to export project table to excel file.'''
        # Ask the location to save the file.
        filename = fd.asksaveasfilename(parent=self, defaultextension='.xls', initialfile='project_details.xls')
        # Checks if a location has been selected, if not will skip the code below.
        if filename != '':
            # This create a temporary workbook and worksheet
            wb = Workbook()
            ws = wb.active

            session = DBSession()
            # Gets the details of project table
            records = session.query(Designation).all()
            # If table is not empty begin writing it into excel worksheet.
            if len(records) != 0:
                ws['A1'] = 'ID'
                ws['B1'] = 'Designation'
                counter = 2
                for record in records:
                    ws[f'A{counter}'] = str(record.id)
                    ws[f'B{counter}'] = str(record.name)
                    counter+=1
            # This will save the workbook from memory into a file.
            wb.save(filename)
            # Then finally closes the connection to the database.
            session.close()

    def close_app(self):
        self.master.destroy()
# End of DesignationWindow class

# Start of AboutWindow class
class AboutWindow(tk.Frame):
    def __init__(self, master=None, *args, **kwargs):
        super(AboutWindow, self).__init__(master, *args, **kwargs)
        self.master.protocol('WM_DELETE_WINDOW', self.close_app)
        self.master.geometry('480x400+20+20')
        self.master.title('About')
        self.img_list = image_list()
        self.img1_list = image_list(size=(48, 48))
        if os.name == "nt":
            self.master.iconbitmap('manpower.ico')
        elif os.name == 'posix':
            self.master.tk.call('wm', 'iconphoto', self.master._w, self.img1_list['manpower'])
        self.style = ttk.Style()
        self.setup_ui()

    def setup_ui(self):
        top_frame = ttk.Frame(self)
        top_frame.pack(fill=tk.X)
        mid_frame = ttk.Frame(self)
        mid_frame.pack(expand=True, fill=tk.BOTH)
        bot_frame = ttk.Frame(self)
        bot_frame.pack(fill=tk.X)

        logo_lbl = ttk.Label(top_frame, image=self.img1_list['manpower'])
        logo_lbl.pack(side=tk.LEFT)
        app_name = ttk.Label(top_frame, text=f'Manpower Management {__version__}')
        app_name.pack(side=tk.LEFT)

        self.noteb = ttk.Notebook(mid_frame)
        self.noteb.pack(expand=True, fill=tk.BOTH)

        close_btn = ttk.Button(bot_frame, text='Close', command=self.close_app)
        close_btn.pack(side=tk.RIGHT)
        close_btn.config(image=self.img_list['cancel'], compound=tk.LEFT)

        details_frame = ttk.Frame(self.noteb)
        self.noteb.add(details_frame, text='Info')

        descp_lbl = ttk.Label(details_frame, text='A simple manpower management software.')
        descp_lbl.pack(padx=10, pady=10)

        web_lbl = ttk.Label(details_frame, text='http://jolazo.c1.biz/manpower')
        web_lbl.pack(padx=10, pady=10)

        copy_lbl = ttk.Label(details_frame, text='Copyright (c) 2022')
        copy_lbl.pack(padx=10, pady=10)

        author_lbl = ttk.Label(details_frame, text='Jesus Vedasto Olazo')
        author_lbl.pack()

        lic_frame = ttk.Frame(self.noteb)
        self.noteb.add(lic_frame, text='License')

        lic_txt = tk.Text(lic_frame, font=('Helvetica', 7, 'normal'))
        lic_txt.pack(expand=True, fill=tk.BOTH)

        with open('LICENSE.txt', 'r') as lfile:
            data = lfile.read()

        lic_txt.insert(tk.END, data)

        close_btn.focus_set()

    def close_app(self):
        self.master.destroy()
# End of AboutWindow class

# Start of TransactionWindow class
class TransactionWindow(tk.Frame):
    def __init__(self, master=None, *args, **kwargs):
        super(TransactionWindow, self).__init__(master, *args, **kwargs)
        self.master.protocol('WM_DELETE_WINDOW', self.close_app)
        # self.master.geometry('600x400+20+20')
        self.master.title('Transaction')
        self.record_id = None
        self.img_list = image_list()
        self.img36_list = image_list(size=(36, 36))
        if os.name == "nt":
            self.master.iconbitmap('manpower.ico')
        elif os.name == 'posix':
            self.master.tk.call('wm', 'iconphoto', self.master._w, self.img36_list['manpower'])
        self.style = ttk.Style()
        self.setup_ui()

    def setup_ui(self):
        top_frame = ttk.Frame(self)
        top_frame.pack(fill=tk.X, padx=5, pady=5)
        mid_frame = ttk.Frame(self)
        mid_frame.pack(expand=True, fill=tk.BOTH, padx=5, pady=5)
        bot_frame = ttk.Frame(self)
        bot_frame.pack(fill=tk.X, padx=5, pady=5)

        proj_name = ttk.Label(top_frame, text='Project Name:')
        proj_name.pack(side=tk.LEFT, padx=5, pady=5)
        self.proj_entry = ttk.Entry(top_frame, font=('Cambria', 10, 'bold'))
        self.proj_entry.pack(side=tk.LEFT, padx=5, pady=5)
        self.search_btn = ttk.Button(top_frame, text='Search', image=self.img_list['search'], compound=tk.LEFT)
        self.search_btn.pack(side=tk.LEFT, padx=5, pady=5)

        self.manp_view = ttk.Treeview(mid_frame)
        self.manp_view.pack(expand=True, fill=tk.BOTH, side=tk.LEFT)
        self.manp_view['columns'] = ('date', 'project', 'present',
                                    'absent', 'vacation', 'total', 'remarks')
        self.manp_view.heading('#0', text='ID')
        self.manp_view.heading('project', text='PROJECT')
        self.manp_view.heading('date', text='DATE')
        self.manp_view.heading('present', text='PRESENT')
        self.manp_view.heading('absent', text='ABSENT')
        self.manp_view.heading('vacation', text='VACATION')
        self.manp_view.heading('total', text='TOTAL')
        self.manp_view.heading('remarks', text='REMARKS')

        self.manp_view.column('#0', width=65, stretch=False)
        self.manp_view.column('date', width=100, stretch=False, anchor=tk.CENTER)
        self.manp_view.column('present', width=70, stretch=False, anchor=tk.CENTER)
        self.manp_view.column('absent', width=70, stretch=False, anchor=tk.CENTER)
        self.manp_view.column('vacation', width=80, stretch=False, anchor=tk.CENTER)
        self.manp_view.column('total', width=70, stretch=False, anchor=tk.CENTER)

        self.manp_view.tag_configure("odd", background="#d5f4e6", font=('Times', 11, ''))
        self.manp_view.tag_configure("even", background="#80ced6", font=('Times', 11, ''))

        self.vbar = ttk.Scrollbar(mid_frame, orient=tk.VERTICAL)
        self.vbar.pack(side=tk.RIGHT, fill=tk.Y)

        self.vbar.config(command=self.manp_view.yview)
        self.manp_view['yscrollcommand'] = self.vbar.set

        self.new_btn = ttk.Button(bot_frame, text='New', image=self.img_list['plus'], compound=tk.LEFT)
        self.new_btn.pack(side=tk.LEFT, padx=5, pady=5)
        self.new_btn.config(command=self.new_record)
        self.edit_btn = ttk.Button(bot_frame, text='Edit', image=self.img_list['edit'], compound=tk.LEFT)
        self.edit_btn.pack(side=tk.LEFT, padx=5, pady=5)
        self.edit_btn.config(command=self.edit_record)
        self.delete_btn = ttk.Button(bot_frame, text='Delete', image=self.img_list['minus'], compound=tk.LEFT)
        self.delete_btn.pack(side=tk.LEFT, padx=5, pady=5)
        self.delete_btn.config(command=self.delete_record)
        self.export_btn = ttk.Button(bot_frame, text='Export', image=self.img_list['export'], compound=tk.LEFT)
        self.export_btn.pack(side=tk.LEFT, padx=5, pady=5)
        self.export_btn.config(command=self.export_record)
        self.close_btn = ttk.Button(bot_frame, text='Close', image=self.img_list['cancel'], compound=tk.LEFT)
        self.close_btn.pack(side=tk.RIGHT, padx=5, pady=5)
        self.close_btn.config(command=self.close_app)

        self.update_view()
        self.proj_entry.focus_set()

    def update_view(self):
        session = DBSession()
        records = session.query(TransactionDetails).join(Transaction).all()
        children = self.manp_view.get_children()
        if len(children) != 0:
            for child in children:
                self.manp_view.delete(child)
        if not len(records) == 0:
            counter = 1
            for record in records:
                tr_date = datetime.strftime(record.transaction.tr_date, '%d/%m/%Y')
                total = record.present + record.absent + record.vacation
                values = (tr_date, record.transaction.project.name, record.present, record.absent, record.vacation, f"{total}", record.transaction.remarks)
                if counter % 2 == 0:
                    self.manp_view.insert('', tk.END, str(record.id), text=str(record.id), tags='even', values=values)
                else:
                    self.manp_view.insert('', tk.END, str(record.id), text=str(record.id), tags='odd', values=values)
                counter+=1
        session.close()

    def new_record(self):
        tp = tk.Toplevel(self)
        win = AddManpowerWindow(tp)
        win.pack(expand=True, fill=tk.BOTH, padx=10, pady=10)
        self.wait_window(tp)
        self.update_view()
        self.proj_entry.focus_set()

    def edit_record(self):
        record = self.manp_view.focus()
        if record == '':
            return
        tp = tk.Toplevel(self)
        win = EditManpowerWindow(tp)
        win.pack(expand=True, fill=tk.BOTH, padx=10, pady=10)
        win.load_record(int(record))
        self.wait_window(tp)
        self.update_view()
        self.proj_entry.focus_set()

    def delete_record(self):
        pass

    def search_record(self):
        pass

    def export_record(self):
        pass

    def close_app(self):
        self.master.destroy()
# End of TransactionWindow class

# Start of AddManpowerWindow class
class AddManpowerWindow(tk.Frame):
    def __init__(self, master=None, *args, **kwargs):
        super(AddManpowerWindow, self).__init__(master, *args, **kwargs)
        self.master.protocol('WM_DELETE_WINDOW', self.close_app)
        # self.master.geometry('600x400+20+20')
        self.master.title('New')
        self.img_list = image_list()
        self.img36_list = image_list(size=(36, 36))
        if os.name == "nt":
            self.master.iconbitmap('manpower.ico')
        elif os.name == 'posix':
            self.master.tk.call('wm', 'iconphoto', self.master._w, self.img36_list['manpower'])
        self.trans_id = None
        self.details_id = None
        self.setup_ui()

    def setup_ui(self):
        top_frame = ttk.Frame(self)
        top_frame.pack(fill=tk.X, padx=5, pady=5)
        mid_frame = ttk.Labelframe(self, text='Details')
        mid_frame.pack(expand=True, fill=tk.BOTH, padx=5, pady=5)
        bot_frame = ttk.Frame(self)
        bot_frame.pack(fill=tk.X, padx=5, pady=5)

        date_lbl = ttk.Label(top_frame, text='Date:')
        date_lbl.grid(row=0, column=0, sticky=tk.W)
        self.date_entry = ttk.Entry(top_frame, width=15)
        self.date_entry.grid(row=0, column=1, sticky=tk.W+tk.E, padx=5, pady=5)
        self.date_entry.insert(tk.END, datetime.strftime(datetime.now(), "%d/%m/%Y"))
        date_btn  = ttk.Button(top_frame, image=self.img_list['calendar'])
        date_btn.grid(row=0, column=2, sticky=tk.W)
        date_btn.config(command=self.change_date)

        self.proj_var = tk.StringVar()
        values = []
        session = DBSession()
        records = session.query(Project.name).all()
        for record in records:
            values.append(record[0])
        session.close()
        self.proj_var.set(values[0])

        proj_lbl = ttk.Label(top_frame, text='Project:')
        proj_lbl.grid(row=1, column=0, sticky=tk.W)
        self.proj_cb = ttk.Combobox(top_frame, textvariable=self.proj_var, values=values)
        self.proj_cb.grid(row=1, column=1, columnspan=2, sticky=tk.W+tk.E, padx=5, pady=5)

        remarks_lbl = ttk.Label(top_frame, text='Remarks:')
        remarks_lbl.grid(row=2, column=0, sticky=tk.W)
        self.remarks_entry = ttk.Entry(top_frame, width=30)
        self.remarks_entry.grid(row=2, column=1, columnspan=2, sticky=tk.W+tk.E, padx=5, pady=5)
        self.trans_save = ttk.Button(top_frame, text='Transaction Save', command=self.save_trans, image=self.img_list['save'], compound=tk.LEFT)
        self.trans_save.grid(row=2, column=3, sticky=tk.W)

        add_detail_frame = ttk.Frame(mid_frame)
        add_detail_frame.pack(fill=tk.X, padx=5, pady=5)
        view_detail_frame = ttk.Frame(mid_frame)
        view_detail_frame.pack(expand=True, fill=tk.BOTH, padx=5, pady=5)

        self.desg_var = tk.StringVar()
        values = []
        session = DBSession()
        records = session.query(Designation.name).all()
        for record in records:
            values.append(record[0])
        self.desg_var.set(values[0])
        session.close()

        self.desg_cb = ttk.Combobox(add_detail_frame, textvariable=self.desg_var, values=values, width=30)
        self.desg_cb.grid(row=1, column=0)

        present_lbl = ttk.Label(add_detail_frame, text='Present')
        present_lbl.grid(row=0, column=1)
        self.present_entry = ttk.Entry(add_detail_frame, width=10, justify=tk.CENTER)
        self.present_entry.grid(row=1, column=1)
        self.present_entry.insert(tk.END, '0')

        absent_lbl = ttk.Label(add_detail_frame, text='Absent')
        absent_lbl.grid(row=0, column=2)
        self.absent_entry = ttk.Entry(add_detail_frame, width=10, justify=tk.CENTER)
        self.absent_entry.grid(row=1, column=2)
        self.absent_entry.insert(tk.END, '0')

        vacation_lbl = ttk.Label(add_detail_frame, text='Vacation')
        vacation_lbl.grid(row=0, column=3)
        self.vacation_entry = ttk.Entry(add_detail_frame, width=10, justify=tk.CENTER)
        self.vacation_entry.grid(row=1, column=3)
        self.vacation_entry.insert(tk.END, '0')

        details_btn = ttk.Button(add_detail_frame, text='Add Entry', image=self.img_list['save'], compound=tk.LEFT, command=self.save_details)
        details_btn.grid(row=1, column=4)

        cols = ('designation', 'present', 'absent', 'vacation', 'total')

        self.details_view = ttk.Treeview(view_detail_frame)
        self.details_view.pack(side=tk.LEFT, expand=True, fill=tk.BOTH)

        self.details_view['columns'] = cols
        self.details_view.column('#0', width=65, stretch=False)
        self.details_view.heading('#0', text='ID')

        for col in cols:
            if col in cols[1:]:
                self.details_view.column(col, width=80, stretch=False)
            self.details_view.heading(col, text=col.upper())

        vbar = ttk.Scrollbar(view_detail_frame, orient=tk.VERTICAL)
        vbar.pack(side=tk.RIGHT, fill=tk.Y)

        vbar.config(command=self.details_view.yview)
        self.details_view['yscrollcommand'] = vbar.set

        self.details_view.tag_configure("odd", background="#d5f4e6")
        self.details_view.tag_configure("even", background="#80ced6")
        self.details_view.tag_configure("total_design", background="#1B1B4B", foreground='#FFFFFF')

        edit_btn = ttk.Button(bot_frame, text='Edit', image=self.img_list['edit'], compound=tk.LEFT)
        edit_btn.pack(side=tk.LEFT)
        edit_btn.config(command=self.edit_record)

        delete_btn = ttk.Button(bot_frame, text='Delete', image=self.img_list['minus'], compound=tk.LEFT)
        delete_btn.pack(side=tk.LEFT)
        delete_btn.config(command=self.delete_record)

        print_btn = ttk.Button(bot_frame, text='Print', image=self.img_list['printer'], compound=tk.LEFT)
        print_btn.pack(side=tk.LEFT)

        export_btn = ttk.Button(bot_frame, text='Export', image=self.img_list['export'], compound=tk.LEFT)
        export_btn.pack(side=tk.LEFT)

        close_btn = ttk.Button(bot_frame, text='Close', image=self.img_list['cancel'], compound=tk.LEFT,command=self.close_app)
        close_btn.pack(side=tk.RIGHT)

        self.date_entry.focus_set()

    def delete_record(self):
        details_id = self.details_view.focus()
        if details_id != '':
            details_id = int(details_id)
            session = DBSession()
            record = session.query(TransactionDetails).filter(TransactionDetails.id == details_id).first()
            session.delete(record)
            session.commit()
            session.close()
            self.update_view()
        else:
            mb.showwarning('No record', 'Please select a record and try again.')

    def save_trans(self):
        tr_date = datetime.strptime(self.date_entry.get(), '%d/%m/%Y')
        project_name = self.proj_var.get()
        remarks = self.remarks_entry.get()
        session = DBSession()
        project = session.query(Project).filter(Project.name == project_name).first()
        new_record = Transaction(tr_date=tr_date, project=project, remarks=remarks)
        session.add(new_record)
        session.commit()
        session.close()
        self.trans_id = self.get_last_id()
        self.trans_save.config(state=tk.DISABLED)

    def get_last_id(self):
        last_id = 0
        session = DBSession()
        record = session.query(Transaction).order_by(Transaction.id.desc()).first()
        if record == None:
            last_id = 1
        else:
            last_id = record.id
        session.close()
        return last_id

    def update_view(self):
        trans_id = self.trans_id
        children = self.details_view.get_children()
        self.details_view.delete(*children)
        session = DBSession()
        records = session.query(TransactionDetails).filter(TransactionDetails.transaction_id == trans_id).all()
        sum_present = 0
        sum_absent = 0
        sum_vacation = 0
        grand_total = 0
        counter = 1
        for record in records:
            sum_present+=record.present
            sum_absent+=record.absent
            sum_vacation+=record.vacation
            values = (record.designation.name, f'{record.present}', f'{record.absent}', f'{record.vacation}', f'{record.present+record.absent+record.vacation}')
            if counter % 2 == 0:
                self.details_view.insert('', tk.END, f'{record.id}', text=f'{record.id}', values=values, tags='even')
            else:
                self.details_view.insert('', tk.END, f'{record.id}', text=f'{record.id}', values=values, tags='odd')
            counter+=1
        grand_total = sum_present + sum_absent + sum_vacation
        values = ('Total', f'{sum_present}', f'{sum_absent}', f'{sum_vacation}', f'{grand_total}')
        self.details_view.insert('', tk.END, 'total' , text='', values=values, tags='total_design')

    def get_details_id(self):
        last_id = 0
        session = DBSession()
        record = session.query(TransactionDetails).order_by(TransactionDetails.id.desc()).first()
        if record == None:
            last_id = 1
        else:
            last_id = record.id
        session.close()
        return last_id

    def save_details(self):
        if self.trans_id == None:
            mb.showwarning('Not save', 'Please save the transaction \nbefore adding details!')
            return

        desg_var = self.desg_var.get()
        present = int(self.present_entry.get())
        absent = int(self.absent_entry.get())
        vacation = int(self.vacation_entry.get())
        session = DBSession()

        if self.details_id != None:
            desg_rec = session.query(Designation).filter(Designation.name == desg_var).first()
            transdetails_rec = session.query(TransactionDetails).filter(TransactionDetails.id == self.details_id).first()
            transdetails_rec.designation = desg_rec
            transdetails_rec.present = present
            transdetails_rec.absent = absent
            transdetails_rec.vacation = vacation
            session.commit()
            session.close()
            self.details_id = None
            self.update_view()
        else:
            trans_rec = session.query(Transaction).filter(Transaction.id == self.trans_id).first()
            desg_rec = session.query(Designation).filter(Designation.name == desg_var).first()
            new_record = TransactionDetails(transaction=trans_rec, designation=desg_rec, present=present, absent=absent, vacation=vacation)
            session.add(new_record)
            session.commit()
            session.close()
            self.update_view()

        # Reset entry quantity to 0
        self.present_entry.delete('0', tk.END)
        self.absent_entry.delete('0', tk.END)
        self.vacation_entry.delete('0', tk.END)
        self.present_entry.insert(tk.END, '0')
        self.absent_entry.insert(tk.END, '0')
        self.vacation_entry.insert(tk.END, '0')
        self.present_entry.focus_set()

    def change_date(self):
        date_now = datetime.today()
        tp = tk.Toplevel(self)
        win = CalendarWidget(date_now.year, date_now.month, tp)
        win.pack(expand=True, fill=tk.BOTH, padx=10, pady=10)
        self.wait_window(tp)
        if win.date is not None:
            self.date_entry.delete('0', tk.END)
            self.date_entry.insert(tk.END, win.date)

    def edit_record(self):
        details_id = self.details_view.focus()
        if details_id != '':
            details_id = int(details_id)
            session = DBSession()
            record = session.query(TransactionDetails).filter(TransactionDetails.id == details_id).first()
            self.desg_var.set(record.designation.name)
            self.present_entry.delete('0', tk.END)
            self.absent_entry.delete('0', tk.END)
            self.vacation_entry.delete('0', tk.END)
            self.present_entry.insert(tk.END, f"{record.present}")
            self.absent_entry.insert(tk.END, f"{record.absent}")
            self.vacation_entry.insert(tk.END, f"{record.vacation}")
            session.close()
            self.details_id = details_id
        else:
            mb.showwarning('No record', 'Please select a record and try again.')

    def close_app(self):
        self.master.destroy()
# End of AddManpowerWindow class

# Start of CalendarWidget class
class CalendarWidget(tk.Frame):

    def __init__(self, year, month, master=None, **kws):
        super(CalendarWidget, self).__init__(master, **kws)
        self.year = year
        self.month = month
        self.date = None
        self.cal = cl.Calendar(6)
        self.master.protocol("WM_DELETE_WINDOW", self.close)
        self.master.title("Calendar")
        self.master.resizable(False, False)
        self.img36_list = image_list(size=(36, 36))
        if os.name == "nt":
            self.master.iconbitmap('manpower.ico')
        elif os.name == 'posix':
            self.master.tk.call('wm', 'iconphoto', self.master._w, self.img36_list['manpower'])
        self.setup_ui()

    def setup_ui(self):
        container = tk.Frame(self)
        container.pack(expand=True, fill=tk.BOTH)

        self.prev_btn = tk.Button(container, text="<",
                                  font="Times 12 bold", fg="blue", bg="white")
        self.prev_btn.pack(side=tk.LEFT, fill="y")
        self.prev_btn.bind("<Button-1>", self.btnHandler)

        self.month_var = tk.StringVar()
        self.month_lbl = tk.Label(container, textvariable=self.month_var,
                                  font="Times 14 bold", fg="blue", bg="white")
        self.month_lbl.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        self.next_btn = tk.Button(container, text=">",
                                  font="Times 12 bold", fg="blue", bg="white")
        self.next_btn.pack(side=tk.LEFT, fill="y")
        self.next_btn.bind("<Button-1>", self.btnHandler)

        self.days_frame = tk.Frame(self)
        self.days_frame.pack(expand=True, fill=tk.BOTH)

        self.updateCalendar()
        self.focus_set()

    def updateCalendar(self):
        children = self.days_frame.winfo_children()
        if len(children) != 0:
            for child in children:
                child.destroy()

        days = ["Sun", "Mon", "Tue",
                "Wed", "Thu", "Fri",
                "Sat"]
        month = ["January", "February", "March",
                 "April", "May", "June",
                 "July", "August", "September",
                 "October", "November", "December"]

        self.month_var.set(month[self.month-1]+" "+str(self.year))
        list_of_days = self.cal.monthdayscalendar(self.year, self.month)
        list_of_days.insert(0, days)

        for idx, week in enumerate(list_of_days):
            for idx1, day in enumerate(week):
                if day == 0:
                    self.day_lbl = tk.Label(self.days_frame, text=str(""),
                                            relief=tk.RAISED,
                                            state="disabled",
                                            font="Times 12 normal",
                                            bg="white")
                    self.day_lbl.grid(row=idx, column=idx1, sticky="nesw")
                elif day in days:
                    self.day_lbl = tk.Label(self.days_frame,
                                            text=str(day), relief=tk.RAISED,
                                            font="Times 12 bold")
                    self.day_lbl.grid(row=idx, column=idx1, sticky="nesw")
                    if (day == "Sun") or (day == "Sat"):
                        self.day_lbl.config(fg="red")
                else:
                    self.day_btn = tk.Button(self.days_frame, text=str(day),
                                             width=3, font="Times 12 normal",
                                             bg="white")
                    self.day_btn.grid(row=idx, column=idx1, sticky="nesw")
                    self.day_btn.bind("<Button-1>", self.printEvent)
                    if (idx1 == 0) or (idx1 == 6):
                        self.day_btn.config(fg='red')

    def printEvent(self, event):
        if event.widget.winfo_class() == "Button":
            day = event.widget.cget("text")
            self.date = "%s/%s/%s" % ("{:02}".format(int(day)),
                                      "{:02}".format(self.month),
                                      str(self.year))
            self.close()

    def btnHandler(self, event):
        if event.widget.cget('text') == "<":
            if (self.month - 1) == 0:
                self.month = 12
                self.year = self.year - 1
                self.updateCalendar()
            else:
                self.month = self.month - 1
                self.updateCalendar()
        elif event.widget.cget('text') == ">":
            if (self.month + 1) > 12:
                self.month = 1
                self.year = self.year + 1
                self.updateCalendar()
            else:
                self.month = self.month + 1
                self.updateCalendar()

    def close(self):
        self.master.destroy()
# End of CalendarWidget class

def main():
    root = tk.Tk()
    win = MainWindow(root)
    win.pack(expand=True, fill=tk.BOTH, padx=5, pady=5)
    root.mainloop()

if __name__ == '__main__':
    sys.exit(main())
